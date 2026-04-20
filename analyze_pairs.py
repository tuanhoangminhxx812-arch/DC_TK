"""Phân tích pattern khớp nhóm mà user phát hiện."""
import pandas as pd
import openpyxl

wb = openpyxl.load_workbook(r"D:\HOC A.I\đối chiếu trung gian\GL_016_33193.xlsx", data_only=True)
ws = wb.active
data_start = data_end = None
for i, row in enumerate(ws.iter_rows(min_row=1), 1):
    vals = [c.value for c in row]
    v0 = str(vals[0]).lower() if vals[0] else ""
    if "dư đầu kỳ" in v0: data_start = i + 1
    if "cộng phát sinh" in v0: data_end = i - 1; break

data_rows = []
for i, row in enumerate(ws.iter_rows(min_row=data_start, max_row=data_end), data_start):
    vals = [c.value for c in row]
    if vals[0] is None and vals[1] is None and vals[5] is None and vals[6] is None: continue
    data_rows.append({
        "STT": len(data_rows)+1,
        "Nguon": vals[0] or "",
        "Ngay": str(vals[1])[:10] if vals[1] else "",
        "Dien giai": vals[4] or "",
        "No": vals[5] if vals[5] is not None else 0,
        "Co": vals[6] if vals[6] is not None else 0,
        "Nguoi lap": str(vals[7]).strip() if vals[7] else "",
    })
df = pd.DataFrame(data_rows)

# === Kiểm tra case cụ thể ===
print("=== Case user chỉ ra ===")
no_stts = [388, 389, 390, 611, 612, 613, 614, 615, 616, 617, 618, 619, 621]
co_stts = [632, 634, 635]

no_df = df[df["STT"].isin(no_stts)]
co_df = df[df["STT"].isin(co_stts)]

print("\nCÁC DÒNG NỢ:")
for _, r in no_df.iterrows():
    print(f"  STT {r['STT']:>3} | {r['Nguoi lap']:<22} | Nợ: {r['No']:>15,.0f} | {r['Dien giai'][:80]}")
sum_no = no_df["No"].sum()
print(f"  TỔNG NỢ: {sum_no:,.0f}")

print("\nCÁC DÒNG CÓ:")
for _, r in co_df.iterrows():
    print(f"  STT {r['STT']:>3} | {r['Nguoi lap']:<22} | Có: {r['Co']:>15,.0f} | {r['Dien giai'][:80]}")
sum_co = co_df["Co"].sum()
print(f"  TỔNG CÓ: {sum_co:,.0f}")

print(f"\n  NỢ - CÓ = {sum_no - sum_co:,.0f}")
print(f"  KHỚP: {'CÓ' if abs(sum_no - sum_co) < 0.01 else 'KHÔNG'}")

# === Xem tất cả 23 dòng chưa khớp còn lại ===
print("\n\n=== TẤT CẢ CÁC DÒNG CHƯA KHỚP (sau phân tích trước) ===")
# Chạy lại phân tích để lấy danh sách chưa khớp
unbalanced_persons = []
for person, gdf in df.groupby("Nguoi lap", sort=False):
    cl = gdf["No"].sum() - gdf["Co"].sum()
    if abs(cl) >= 0.01: unbalanced_persons.append(person)

global_matched = set()
for person in unbalanced_persons:
    pdf = df[df["Nguoi lap"]==person].copy().reset_index(drop=True)
    matched = [False]*len(pdf)
    no_rows, co_rows = [], []
    for idx, row in pdf.iterrows():
        nv = row["No"] if pd.notna(row["No"]) else 0
        cv = row["Co"] if pd.notna(row["Co"]) else 0
        if abs(nv)>0 and abs(cv)==0: no_rows.append(idx)
        elif abs(cv)>0 and abs(nv)==0: co_rows.append(idx)
    used_co = set()
    for ni in no_rows:
        if matched[ni]: continue
        nv = pdf.loc[ni,"No"]
        for ci in co_rows:
            if ci in used_co or matched[ci]: continue
            cv = pdf.loc[ci,"Co"]
            if abs(abs(nv)-abs(cv))<0.01: matched[ni]=matched[ci]=True; used_co.add(ci); break
    umi = [i for i in no_rows if not matched[i]]
    pos_no = [i for i in umi if pdf.loc[i,"No"]>0]
    neg_no = [i for i in umi if pdf.loc[i,"No"]<0]
    used_neg = set()
    for pi in pos_no:
        if matched[pi]: continue
        pv=pdf.loc[pi,"No"]
        for ngi in neg_no:
            if ngi in used_neg or matched[ngi]: continue
            nv=pdf.loc[ngi,"No"]
            if abs(pv+nv)<0.01: matched[pi]=matched[ngi]=True; used_neg.add(ngi); break
    uci = [i for i in co_rows if not matched[i]]
    pos_co = [i for i in uci if pdf.loc[i,"Co"]>0]
    neg_co = [i for i in uci if pdf.loc[i,"Co"]<0]
    used_neg_co = set()
    for pi in pos_co:
        if matched[pi]: continue
        pv=pdf.loc[pi,"Co"]
        for ngi in neg_co:
            if ngi in used_neg_co or matched[ngi]: continue
            nv=pdf.loc[ngi,"Co"]
            if abs(pv+nv)<0.01: matched[pi]=matched[ngi]=True; used_neg_co.add(ngi); break
    for idx in range(len(pdf)):
        if matched[idx]: global_matched.add(pdf.loc[idx,"STT"])

# Cross match 1-1
unmatched_df = df[(df["Nguoi lap"].isin(unbalanced_persons))&(~df["STT"].isin(global_matched))].copy()
cross_no = unmatched_df[(unmatched_df["No"].abs()>0)&((unmatched_df["Co"].fillna(0)).abs()==0)]
cross_co = unmatched_df[(unmatched_df["Co"].abs()>0)&((unmatched_df["No"].fillna(0)).abs()==0)]
cross_stts = set()
used = set()
for _, nr in cross_no.iterrows():
    if nr["STT"] in cross_stts: continue
    for _, cr in cross_co.iterrows():
        if cr["STT"] in used or cr["STT"] in cross_stts: continue
        if abs(abs(nr["No"])-abs(cr["Co"]))<0.01:
            cross_stts.add(nr["STT"]); cross_stts.add(cr["STT"]); used.add(cr["STT"]); break
global_matched.update(cross_stts)

# In các dòng còn lại
remaining = df[(df["Nguoi lap"].isin(unbalanced_persons))&(~df["STT"].isin(global_matched))]
print(f"\nTổng còn lại: {len(remaining)} dòng\n")

for person in remaining["Nguoi lap"].unique():
    pr = remaining[remaining["Nguoi lap"]==person]
    sn = pr["No"].sum()
    sc = pr["Co"].sum()
    print(f"\n--- {person} ({len(pr)} dòng) | Nợ: {sn:,.0f} | Có: {sc:,.0f} | CL: {sn-sc:,.0f} ---")
    for _, r in pr.iterrows():
        no_str = f"Nợ: {r['No']:>15,.0f}" if r["No"] != 0 else "                    "
        co_str = f"Có: {r['Co']:>15,.0f}" if r["Co"] != 0 else "                    "
        print(f"  STT {r['STT']:>3} | {no_str} | {co_str} | {r['Dien giai'][:70]}")
