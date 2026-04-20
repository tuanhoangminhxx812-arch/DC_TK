import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime
from difflib import SequenceMatcher
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from copy import copy

# ─── Page Config ───
st.set_page_config(
    page_title="Đối Chiếu Trung Gian",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─── Custom CSS ───
st.markdown("""
<style>
/* ── Import Google Fonts ── */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

/* ── Global ── */
html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
}

/* ── Main container ── */
.main .block-container {
    padding-top: 1.5rem;
    padding-bottom: 1rem;
    max-width: 1400px;
}

/* ── Hero Header ── */
.hero-header {
    background: linear-gradient(135deg, #0f0c29 0%, #302b63 50%, #24243e 100%);
    border-radius: 20px;
    padding: 2rem 2.5rem;
    margin-bottom: 1.5rem;
    box-shadow: 0 20px 60px rgba(0,0,0,0.3);
    position: relative;
    overflow: hidden;
}
.hero-header::before {
    content: '';
    position: absolute;
    top: -50%;
    right: -20%;
    width: 400px;
    height: 400px;
    background: radial-gradient(circle, rgba(99,102,241,0.15) 0%, transparent 70%);
    border-radius: 50%;
}
.hero-header h1 {
    color: #fff;
    font-size: 2rem;
    font-weight: 800;
    margin: 0;
    letter-spacing: -0.5px;
    background: linear-gradient(90deg, #e0e7ff, #a5b4fc, #818cf8);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
}
.hero-header p {
    color: rgba(255,255,255,0.6);
    font-size: 0.95rem;
    margin-top: 0.4rem;
    font-weight: 400;
}

/* ── Stat Cards ── */
.stat-card {
    background: rgba(255,255,255,0.03);
    backdrop-filter: blur(10px);
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 16px;
    padding: 1.3rem 1.5rem;
    text-align: center;
    transition: all 0.3s ease;
}
.stat-card:hover {
    transform: translateY(-2px);
    border-color: rgba(129,140,248,0.3);
    box-shadow: 0 8px 30px rgba(99,102,241,0.15);
}
.stat-label {
    color: rgba(255,255,255,0.5);
    font-size: 0.75rem;
    text-transform: uppercase;
    letter-spacing: 1px;
    font-weight: 600;
}
.stat-value {
    font-size: 1.5rem;
    font-weight: 700;
    margin-top: 0.3rem;
}
.stat-value.blue { color: #818cf8; }
.stat-value.green { color: #34d399; }
.stat-value.red { color: #f87171; }
.stat-value.amber { color: #fbbf24; }

/* ── Glass panel ── */
.glass-panel {
    background: rgba(255,255,255,0.02);
    backdrop-filter: blur(12px);
    border: 1px solid rgba(255,255,255,0.06);
    border-radius: 16px;
    padding: 1.5rem;
    margin-bottom: 1rem;
}

/* ── Status badges ── */
.badge-ok {
    display: inline-block;
    background: linear-gradient(135deg, #059669, #34d399);
    color: #fff;
    padding: 0.4rem 1.2rem;
    border-radius: 30px;
    font-weight: 600;
    font-size: 0.9rem;
    box-shadow: 0 4px 15px rgba(52,211,153,0.3);
}
.badge-error {
    display: inline-block;
    background: linear-gradient(135deg, #dc2626, #f87171);
    color: #fff;
    padding: 0.4rem 1.2rem;
    border-radius: 30px;
    font-weight: 600;
    font-size: 0.9rem;
    box-shadow: 0 4px 15px rgba(248,113,113,0.3);
}

/* ── Difference row table ── */
.diff-table {
    width: 100%;
    border-collapse: separate;
    border-spacing: 0;
    border-radius: 12px;
    overflow: hidden;
    margin-top: 1rem;
}
.diff-table thead th {
    background: linear-gradient(135deg, #312e81, #4338ca);
    color: #e0e7ff;
    padding: 0.8rem 1rem;
    font-weight: 600;
    font-size: 0.8rem;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    border: none;
}
.diff-table tbody td {
    padding: 0.65rem 1rem;
    border-bottom: 1px solid rgba(255,255,255,0.04);
    color: rgba(255,255,255,0.85);
    font-size: 0.85rem;
}
.diff-table tbody tr {
    transition: all 0.2s ease;
}
.diff-table tbody tr:hover {
    background: rgba(99,102,241,0.08);
}
.diff-table tbody tr:nth-child(even) {
    background: rgba(255,255,255,0.02);
}

/* ── Buttons ── */
.stButton > button {
    border-radius: 12px;
    font-weight: 600;
    font-family: 'Inter', sans-serif;
    padding: 0.5rem 1.5rem;
    transition: all 0.3s ease;
    border: none;
}

/* ── File uploader ── */
[data-testid="stFileUploader"] {
    border: 2px dashed rgba(129,140,248,0.3);
    border-radius: 16px;
    padding: 1rem;
    transition: all 0.3s ease;
}
[data-testid="stFileUploader"]:hover {
    border-color: rgba(129,140,248,0.6);
    background: rgba(99,102,241,0.05);
}

/* ── Dataframe ── */
[data-testid="stDataFrame"] {
    border-radius: 12px;
    overflow: hidden;
}

/* ── Download button ── */
.stDownloadButton > button {
    background: linear-gradient(135deg, #059669, #34d399) !important;
    color: #fff !important;
    border: none !important;
    border-radius: 12px !important;
    font-weight: 600 !important;
    padding: 0.6rem 2rem !important;
    box-shadow: 0 4px 15px rgba(52,211,153,0.25) !important;
}
.stDownloadButton > button:hover {
    box-shadow: 0 8px 25px rgba(52,211,153,0.4) !important;
    transform: translateY(-1px);
}

/* ── Animations ── */
@keyframes fadeInUp {
    from { opacity: 0; transform: translateY(20px); }
    to { opacity: 1; transform: translateY(0); }
}
.animate-in {
    animation: fadeInUp 0.5s ease forwards;
}

/* ── Divider ── */
.section-divider {
    height: 1px;
    background: linear-gradient(90deg, transparent, rgba(129,140,248,0.3), transparent);
    margin: 1.5rem 0;
}
</style>
""", unsafe_allow_html=True)


def format_number(val):
    """Format number with thousands separator"""
    if pd.isna(val) or val is None:
        return ""
    try:
        num = float(val)
        if num == 0:
            return "0"
        if num == int(num):
            return f"{int(num):,}".replace(",", ".")
        return f"{num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return str(val)


def parse_excel_file(uploaded_file):
    """Parse the uploaded Excel file and extract data rows.
    Supports both old format (GL_016) and new format (0903).
    Automatically searches all sheets to find the proper data table.
    """
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    
    ws = None
    is_new_format = False
    header_row = None
    data_start = None
    data_end = None

    # ── Search through all worksheets to find the data structure ──
    for sheet in wb.worksheets:
        temp_is_new_format = False
        temp_header_row = None
        temp_data_start = None
        temp_data_end = None
        
        # Scan the first 100 rows to find headers
        for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=100), 1):
            vals = [cell.value for cell in row]
            row_text = ' '.join([str(v).lower().strip() for v in vals[:10] if v is not None])

            # New format detection (0903)
            if 'nguồn phát sinh' in row_text:
                temp_is_new_format = True
                temp_header_row = i
                if temp_data_start is None:
                    temp_data_start = i + 1
            # Old format detection (GL_016)
            elif 'nguồn bút toán' in row_text:
                temp_header_row = i
                
            # Old format: data start after "Số dư đầu kỳ"
            if 'số dư đầu kỳ' in row_text:
                temp_data_start = i + 1
                
            # End markers
            if 'tổng cộng' in row_text or 'cộng phát sinh' in row_text:
                temp_data_end = i - 1
                break
                
        # If we found a valid starting point, we select this sheet
        if temp_data_start is not None:
            ws = sheet
            is_new_format = temp_is_new_format
            header_row = temp_header_row
            data_start = temp_data_start
            data_end = temp_data_end
            
            # If the end marker wasn't found in the first 100 rows, scan the rest of the sheet
            if data_end is None:
                for i, row in enumerate(ws.iter_rows(min_row=data_start, max_row=ws.max_row), data_start):
                    vals = [cell.value for cell in row]
                    row_text = ' '.join([str(v).lower().strip() for v in vals[:10] if v is not None])
                    if 'tổng cộng' in row_text or 'cộng phát sinh' in row_text:
                        data_end = i - 1
                        break
            break

    # If no sheet matches the expected structure
    if ws is None or data_start is None:
        if ws is None: ws = wb.active # fallback
        st.error("Không tìm thấy cấu trúc dữ liệu phù hợp trong file!")
        return None, None, None, None

    if data_end is None:
        data_end = ws.max_row
        
    sheet_name = ws.title

    # ── Extract data rows ──
    data_rows = []
    for i, row in enumerate(ws.iter_rows(min_row=data_start, max_row=data_end), data_start):
        vals = [cell.value for cell in row]
        vals.extend([None] * (25 - len(vals)))  # Padding to avoid IndexError

        if is_new_format:
            # New format (0903): 21 columns
            if vals[0] is None and vals[1] is None and vals[13] is None and vals[14] is None:
                continue
            data_rows.append({
                'STT': len(data_rows) + 1,
                'Dòng Excel': i,
                'Nguồn phát sinh': vals[0] if vals[0] else '',
                'Số giao dịch': str(vals[1]).lstrip("'") if vals[1] else '',
                'Ngày giao dịch': vals[2].strftime('%d/%m/%Y') if hasattr(vals[2], 'strftime') else (str(vals[2]) if vals[2] else ''),
                'Nội dung': vals[17] if vals[17] else '',
                'Nợ nguyên tệ': vals[13] if vals[13] is not None else 0,
                'Có nguyên tệ': vals[14] if vals[14] is not None else 0,
                'Người tạo': str(vals[19]).strip() if vals[19] else '',
            })
        else:
            # Old format (GL_016): 8 columns
            if vals[0] is None and vals[1] is None and vals[5] is None and vals[6] is None:
                continue
            data_rows.append({
                'STT': len(data_rows) + 1,
                'Dòng Excel': i,
                'Nguồn phát sinh': vals[0] if vals[0] else '',
                'Số giao dịch': str(vals[2]).lstrip("'") if vals[2] else '',
                'Ngày giao dịch': vals[1].strftime('%d/%m/%Y') if hasattr(vals[1], 'strftime') else (str(vals[1]) if vals[1] else ''),
                'Nội dung': vals[4] if vals[4] else '',
                'Nợ nguyên tệ': vals[5] if vals[5] is not None else 0,
                'Có nguyên tệ': vals[6] if vals[6] is not None else 0,
                'Người tạo': str(vals[7]).strip() if vals[7] else '',
            })

    df = pd.DataFrame(data_rows)

    # ── Get totals from the file ──
    total_vals = None
    for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row):
        vals = [cell.value for cell in row]
        row_text = ' '.join([str(v).lower().strip() for v in vals[:10] if v is not None])
        if 'tổng cộng' in row_text or 'cộng phát sinh' in row_text:
            vals.extend([None] * (25 - len(vals)))
            if is_new_format:
                total_vals = {
                    'Tổng Nợ (file)': vals[13] if vals[13] else 0,
                    'Tổng Có (file)': vals[14] if vals[14] else 0,
                }
            else:
                total_vals = {
                    'Tổng Nợ (file)': vals[5] if vals[5] else 0,
                    'Tổng Có (file)': vals[6] if vals[6] else 0,
                }
            break

    return df, total_vals, wb, sheet_name


def text_similarity(text1, text2):
    """Calculate similarity ratio between two texts (0.0 to 1.0)."""
    if not text1 or not text2:
        return 0.0
    s1 = str(text1).strip().lower()
    s2 = str(text2).strip().lower()
    if s1 == s2:
        return 1.0
    return SequenceMatcher(None, s1, s2).ratio()


def analyze_data(df, total_vals):
    """Analyze debit vs credit and find discrepancies using smart matching."""

    # Calculate totals from data
    total_no = df['Nợ nguyên tệ'].sum()
    total_co = df['Có nguyên tệ'].sum()
    chenh_lech = total_no - total_co

    results = {
        'total_no': total_no,
        'total_co': total_co,
        'chenh_lech': chenh_lech,
        'total_no_file': total_vals['Tổng Nợ (file)'] if total_vals else 0,
        'total_co_file': total_vals['Tổng Có (file)'] if total_vals else 0,
        'is_balanced': abs(chenh_lech) < 0.01,
    }

    # ════════════════════════════════════════════════
    # Step 1: Group by "Người lập" and check balance
    # ════════════════════════════════════════════════
    person_groups = df.groupby('Người tạo', sort=False)
    person_summary = []
    unbalanced_persons = []

    for person, group_df in person_groups:
        sum_no = group_df['Nợ nguyên tệ'].sum()
        sum_co = group_df['Có nguyên tệ'].sum()
        cl = sum_no - sum_co
        person_summary.append({
            'Người tạo': person,
            'Tổng Nợ': sum_no,
            'Tổng Có': sum_co,
            'Chênh lệch': cl,
            'Số dòng': len(group_df),
            'Cân bằng': abs(cl) < 0.01,
        })
        if abs(cl) >= 0.01:
            unbalanced_persons.append(person)

    person_summary_df = pd.DataFrame(person_summary)

    # ════════════════════════════════════════════════
    # Step 2: For unbalanced persons, do smart matching
    # ════════════════════════════════════════════════
    # Dùng global index để theo dõi matched trên toàn bộ df
    global_matched = set()  # Lưu STT đã khớp
    matched_details = {}

    for person in unbalanced_persons:
        person_df = df[df['Người tạo'] == person].copy()
        person_df = person_df.reset_index(drop=True)

        matched = [False] * len(person_df)

        # Phân loại dòng Nợ và Có
        no_rows = []
        co_rows = []
        for idx, row in person_df.iterrows():
            no_val = row['Nợ nguyên tệ'] if pd.notna(row['Nợ nguyên tệ']) else 0
            co_val = row['Có nguyên tệ'] if pd.notna(row['Có nguyên tệ']) else 0
            if abs(no_val) > 0 and abs(co_val) == 0:
                no_rows.append(idx)
            elif abs(co_val) > 0 and abs(no_val) == 0:
                co_rows.append(idx)

        # ──── Pass 1: Khớp Nợ = Có cùng người lập (chỉ cần số tiền bằng nhau) ────
        used_co = set()
        for ni in no_rows:
            if matched[ni]:
                continue
            no_val = person_df.loc[ni, 'Nợ nguyên tệ']

            for ci in co_rows:
                if ci in used_co or matched[ci]:
                    continue
                co_val = person_df.loc[ci, 'Có nguyên tệ']

                if abs(abs(no_val) - abs(co_val)) < 0.01:
                    matched[ni] = True
                    matched[ci] = True
                    used_co.add(ci)
                    break

        # ──── Pass 2: Nợ dương + Nợ âm triệt tiêu (chỉ cần số tiền) ────
        unmatched_no_indices = [i for i in no_rows if not matched[i]]
        positive_no = [i for i in unmatched_no_indices if person_df.loc[i, 'Nợ nguyên tệ'] > 0]
        negative_no = [i for i in unmatched_no_indices if person_df.loc[i, 'Nợ nguyên tệ'] < 0]

        used_neg = set()
        for pi in positive_no:
            if matched[pi]:
                continue
            p_val = person_df.loc[pi, 'Nợ nguyên tệ']

            for ngi in negative_no:
                if ngi in used_neg or matched[ngi]:
                    continue
                n_val = person_df.loc[ngi, 'Nợ nguyên tệ']

                if abs(p_val + n_val) < 0.01:
                    matched[pi] = True
                    matched[ngi] = True
                    used_neg.add(ngi)
                    break

        # ──── Pass 2b: Nhóm Nợ triệt tiêu (tổng nhóm = 0) ────
        still_unmatched_no = [i for i in no_rows if not matched[i]]
        if len(still_unmatched_no) > 1:
            no_sum = sum(person_df.loc[i, 'Nợ nguyên tệ'] for i in still_unmatched_no)
            if abs(no_sum) < 0.01:
                for i in still_unmatched_no:
                    matched[i] = True

        # ──── Pass 3: Có dương + Có âm triệt tiêu (chỉ cần số tiền) ────
        unmatched_co_indices = [i for i in co_rows if not matched[i]]
        positive_co = [i for i in unmatched_co_indices if person_df.loc[i, 'Có nguyên tệ'] > 0]
        negative_co = [i for i in unmatched_co_indices if person_df.loc[i, 'Có nguyên tệ'] < 0]

        used_neg_co = set()
        for pi in positive_co:
            if matched[pi]:
                continue
            p_val = person_df.loc[pi, 'Có nguyên tệ']

            for ngi in negative_co:
                if ngi in used_neg_co or matched[ngi]:
                    continue
                n_val = person_df.loc[ngi, 'Có nguyên tệ']

                if abs(p_val + n_val) < 0.01:
                    matched[pi] = True
                    matched[ngi] = True
                    used_neg_co.add(ngi)
                    break

        # ──── Pass 3b: Nhóm Có triệt tiêu (tổng nhóm = 0) ────
        still_unmatched_co = [i for i in co_rows if not matched[i]]
        if len(still_unmatched_co) > 1:
            co_sum = sum(person_df.loc[i, 'Có nguyên tệ'] for i in still_unmatched_co)
            if abs(co_sum) < 0.01:
                for i in still_unmatched_co:
                    matched[i] = True

        # Ghi nhận kết quả matching cho người này
        matched_count = sum(matched)
        unmatched_count = len(person_df) - matched_count
        matched_details[person] = {
            'total_rows': len(person_df),
            'matched': matched_count,
            'unmatched': unmatched_count,
        }

        # Ghi nhận STT đã khớp vào global
        for idx in range(len(person_df)):
            if matched[idx]:
                global_matched.add(person_df.loc[idx, 'STT'])

    # ════════════════════════════════════════════════
    # Step 3 (Pass 4): Khớp CHÉO giữa các người lập
    # Nợ người A = Có người B (cùng số tiền)
    # ════════════════════════════════════════════════
    # Gom tất cả dòng chưa khớp từ các người lập có chênh lệch
    unmatched_df = df[
        (df['Người tạo'].isin(unbalanced_persons)) &
        (~df['STT'].isin(global_matched))
    ].copy().reset_index(drop=True)

    cross_matched_stts = set()

    if len(unmatched_df) > 0:
        # Lấy dòng có Nợ và dòng có Có
        cross_no = unmatched_df[
            (unmatched_df['Nợ nguyên tệ'].abs() > 0) &
            ((unmatched_df['Có nguyên tệ'].fillna(0)).abs() == 0)
        ]
        cross_co = unmatched_df[
            (unmatched_df['Có nguyên tệ'].abs() > 0) &
            ((unmatched_df['Nợ nguyên tệ'].fillna(0)).abs() == 0)
        ]

        used_cross_co = set()
        for _, no_row in cross_no.iterrows():
            if no_row['STT'] in cross_matched_stts:
                continue
            no_val = no_row['Nợ nguyên tệ']

            for _, co_row in cross_co.iterrows():
                if co_row['STT'] in used_cross_co or co_row['STT'] in cross_matched_stts:
                    continue
                co_val = co_row['Có nguyên tệ']

                # Cùng số tiền (khác người lập)
                if abs(abs(no_val) - abs(co_val)) < 0.01:
                    cross_matched_stts.add(no_row['STT'])
                    cross_matched_stts.add(co_row['STT'])
                    used_cross_co.add(co_row['STT'])
                    break

    # Cập nhật global_matched
    global_matched.update(cross_matched_stts)

    # ════════════════════════════════════════════════
    # Step 3b (Pass 5): Khớp NHÓM — tổng nhiều dòng Nợ = tổng nhiều dòng Có
    # ════════════════════════════════════════════════
    remaining_df = df[
        (df['Người tạo'].isin(unbalanced_persons)) &
        (~df['STT'].isin(global_matched))
    ].copy().reset_index(drop=True)

    group_matched_stts = set()

    if len(remaining_df) > 0:
        # Tách dòng Nợ và Có chưa khớp
        rem_no = remaining_df[
            (remaining_df['Nợ nguyên tệ'].fillna(0).abs() > 0)
        ].copy()
        rem_co = remaining_df[
            (remaining_df['Có nguyên tệ'].fillna(0).abs() > 0) &
            (remaining_df['Nợ nguyên tệ'].fillna(0).abs() == 0)
        ].copy()

        # Net value cho mỗi dòng (Nợ dương, Có âm)
        no_items = []  # (STT, value)
        for _, r in rem_no.iterrows():
            nv = r['Nợ nguyên tệ'] if pd.notna(r['Nợ nguyên tệ']) else 0
            if abs(nv) > 0:
                no_items.append((r['STT'], nv))

        co_items = []
        for _, r in rem_co.iterrows():
            cv = r['Có nguyên tệ'] if pd.notna(r['Có nguyên tệ']) else 0
            if abs(cv) > 0:
                co_items.append((r['STT'], cv))

        # Chiến lược: tìm nhóm Nợ có tổng = nhóm Có có tổng
        # Dùng cách tiếp cận: thử từng tổ hợp Có, tìm tổ hợp Nợ khớp
        # Giới hạn: chỉ thử nhóm <= 20 dòng để tránh quá chậm

        if no_items and co_items:
            # Tính tổng Nợ chưa khớp và tổng Có chưa khớp
            total_rem_no = sum(v for _, v in no_items)
            total_rem_co = sum(v for _, v in co_items)

            # Nếu tổng Nợ == tổng Có → tất cả khớp nhau
            if abs(total_rem_no - total_rem_co) < 0.01:
                for stt, _ in no_items:
                    group_matched_stts.add(stt)
                for stt, _ in co_items:
                    group_matched_stts.add(stt)
            else:
                # Thử tìm subset: Nhóm Nợ nào có tổng = Nhóm Có nào
                # Dùng dict lưu tổng → danh sách STT cho phía Có
                from itertools import combinations

                # Giới hạn cực kỳ chặt chẽ để tránh bùng nổ tổ hợp
                if len(co_items) > 30: max_co_combo = 2
                elif len(co_items) > 15: max_co_combo = 3
                else: max_co_combo = min(len(co_items), 5)
                
                co_sum_map = {}  # sum -> list of STTs

                for size in range(1, max_co_combo + 1):
                    for combo in combinations(co_items, size):
                        combo_sum = sum(v for _, v in combo)
                        combo_stts = frozenset(s for s, _ in combo)
                        # Lưu nhiều combo cùng sum
                        rounded_sum = round(combo_sum, 2)
                        if rounded_sum not in co_sum_map:
                            co_sum_map[rounded_sum] = []
                        co_sum_map[rounded_sum].append(combo_stts)

                # Tìm subset Nợ khớp
                if len(no_items) > 30: max_no_combo = 2
                elif len(no_items) > 15: max_no_combo = 3
                else: max_no_combo = min(len(no_items), 5)
                
                found_groups = []

                for size in range(1, max_no_combo + 1):
                    for combo in combinations(no_items, size):
                        combo_sum = round(sum(v for _, v in combo), 2)
                        if combo_sum in co_sum_map:
                            no_stts = frozenset(s for s, _ in combo)
                            # Kiểm tra không trùng với đã khớp
                            if not no_stts.intersection(group_matched_stts):
                                for co_stts in co_sum_map[combo_sum]:
                                    if not co_stts.intersection(group_matched_stts):
                                        found_groups.append((no_stts, co_stts))
                                        group_matched_stts.update(no_stts)
                                        group_matched_stts.update(co_stts)
                                        break
                    if len(group_matched_stts) > 0 and size > 5:
                        # Đã tìm được, kiểm tra còn dòng nào không
                        remaining_no = [x for x in no_items if x[0] not in group_matched_stts]
                        remaining_co = [x for x in co_items if x[0] not in group_matched_stts]
                        if not remaining_no or not remaining_co:
                            break

    global_matched.update(group_matched_stts)

    # Cập nhật matched_details cho cả cross + group matching
    all_new_matched = cross_matched_stts | group_matched_stts
    if all_new_matched:
        cross_count_per_person = {}
        for stt in all_new_matched:
            person = df[df['STT'] == stt]['Người tạo'].values[0]
            cross_count_per_person[person] = cross_count_per_person.get(person, 0) + 1

        for person, count in cross_count_per_person.items():
            if person in matched_details:
                matched_details[person]['matched'] += count
                matched_details[person]['unmatched'] -= count

    # ════════════════════════════════════════════════
    # Step 4: Thu thập dòng cuối cùng chưa khớp
    # ════════════════════════════════════════════════
    all_unmatched_rows = []
    for _, row in df.iterrows():
        if row['Người tạo'] in unbalanced_persons and row['STT'] not in global_matched:
            no_val = row['Nợ nguyên tệ'] if pd.notna(row['Nợ nguyên tệ']) else 0
            co_val = row['Có nguyên tệ'] if pd.notna(row['Có nguyên tệ']) else 0
            row_data = row.to_dict()
            row_data['Chênh lệch (Nợ - Có)'] = no_val - co_val
            all_unmatched_rows.append(row_data)

    discrepancy_df = pd.DataFrame(all_unmatched_rows) if all_unmatched_rows else pd.DataFrame()
    results['person_summary_df'] = person_summary_df
    results['unbalanced_persons'] = unbalanced_persons
    results['matched_details'] = matched_details
    results['cross_matched_count'] = len(cross_matched_stts)

    return results, discrepancy_df


def create_export(df_original, df_discrepancy, results):
    """Create Excel export with 2 sheets."""
    output = BytesIO()
    wb = openpyxl.Workbook()

    # ─── Sheet 1: Original Data ───
    ws1 = wb.active
    ws1.title = "Dữ liệu gốc"

    # Styles
    header_font = Font(name='Times New Roman', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='312e81', end_color='312e81', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='Times New Roman', size=10)
    cell_align = Alignment(vertical='center', wrap_text=True)
    number_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws1.merge_cells('A1:I1')
    ws1['A1'] = 'DỮ LIỆU GỐC - SỔ CHI TIẾT TÀI KHOẢN'
    ws1['A1'].font = Font(name='Times New Roman', bold=True, size=14, color='312e81')
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 35

    # export date
    ws1.merge_cells('A2:I2')
    ws1['A2'] = f'Ngày xuất: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws1['A2'].font = Font(name='Times New Roman', italic=True, size=9, color='666666')
    ws1['A2'].alignment = Alignment(horizontal='center')

    # Headers
    headers = ['STT', 'Dòng Excel', 'Nguồn phát sinh', 'Số giao dịch', 'Ngày giao dịch',
               'Nội dung', 'Nợ nguyên tệ', 'Có nguyên tệ']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws1.row_dimensions[4].height = 30

    # Data
    for i, (_, row) in enumerate(df_original.iterrows(), 5):
        for j, col_name in enumerate(headers, 1):
            val = row.get(col_name, '')
            cell = ws1.cell(row=i, column=j, value=val)
            cell.font = cell_font
            cell.border = thin_border
            if col_name in ('Nợ nguyên tệ', 'Có nguyên tệ'):
                cell.alignment = number_align
                cell.number_format = '#,##0'
            else:
                cell.alignment = cell_align

    # Totals row
    total_row = 5 + len(df_original)
    ws1.merge_cells(f'A{total_row}:F{total_row}')
    ws1.cell(row=total_row, column=1, value='TỔNG CỘNG').font = Font(name='Times New Roman', bold=True, size=11)
    ws1.cell(row=total_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws1.cell(row=total_row, column=1).border = thin_border
    for c in range(2, 7):
        ws1.cell(row=total_row, column=c).border = thin_border

    ws1.cell(row=total_row, column=7, value=results['total_no']).font = Font(name='Times New Roman', bold=True, size=11)
    ws1.cell(row=total_row, column=7).number_format = '#,##0'
    ws1.cell(row=total_row, column=7).alignment = number_align
    ws1.cell(row=total_row, column=7).border = thin_border

    ws1.cell(row=total_row, column=8, value=results['total_co']).font = Font(name='Times New Roman', bold=True, size=11)
    ws1.cell(row=total_row, column=8).number_format = '#,##0'
    ws1.cell(row=total_row, column=8).alignment = number_align
    ws1.cell(row=total_row, column=8).border = thin_border

    # Column widths
    widths = [6, 10, 22, 18, 12, 50, 20, 20]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ─── Sheet 2: Report ───
    ws2 = wb.create_sheet(title="Báo cáo chênh lệch")

    # Title
    ws2.merge_cells('A1:J1')
    ws2['A1'] = 'BÁO CÁO KẾT QUẢ ĐỐI CHIẾU'
    ws2['A1'].font = Font(name='Times New Roman', bold=True, size=14, color='312e81')
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 35

    ws2.merge_cells('A2:J2')
    ws2['A2'] = f'Ngày xuất: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws2['A2'].font = Font(name='Times New Roman', italic=True, size=9, color='666666')
    ws2['A2'].alignment = Alignment(horizontal='center')

    # Summary section
    summary_fill = PatternFill(start_color='f0f0ff', end_color='f0f0ff', fill_type='solid')

    ws2.merge_cells('A4:C4')
    ws2['A4'] = 'TÓM TẮT KẾT QUẢ'
    ws2['A4'].font = Font(name='Times New Roman', bold=True, size=12, color='312e81')

    summary_data = [
        ('Tổng phát sinh Nợ:', results['total_no']),
        ('Tổng phát sinh Có:', results['total_co']),
        ('Chênh lệch (Nợ - Có):', results['chenh_lech']),
        ('Kết luận:', 'CÂN BẰNG ✓' if results['is_balanced'] else 'CHÊNH LỆCH ✗'),
    ]
    for idx, (label, value) in enumerate(summary_data, 5):
        ws2.cell(row=idx, column=1, value=label).font = Font(name='Times New Roman', bold=True, size=11)
        ws2.cell(row=idx, column=1).fill = summary_fill
        ws2.cell(row=idx, column=1).border = thin_border
        ws2.merge_cells(f'A{idx}:B{idx}')
        cell = ws2.cell(row=idx, column=3, value=value)
        cell.font = Font(name='Times New Roman', bold=True, size=11,
                         color='059669' if (idx == 8 and results['is_balanced']) else
                         ('dc2626' if idx == 8 else '000000'))
        if isinstance(value, (int, float)):
            cell.number_format = '#,##0'
        cell.fill = summary_fill
        cell.border = thin_border

    # Discrepancy table
    if not df_discrepancy.empty:
        start_row = 11
        ws2.merge_cells(f'A{start_row}:J{start_row}')
        ws2[f'A{start_row}'] = 'CÁC DÒNG GÂY CHÊNH LỆCH'
        ws2[f'A{start_row}'].font = Font(name='Times New Roman', bold=True, size=12, color='dc2626')
        ws2.row_dimensions[start_row].height = 25

        # Headers
        disc_headers = ['STT', 'Dòng Excel', 'Nguồn phát sinh', 'Số giao dịch', 'Ngày giao dịch',
                        'Nội dung', 'Nợ nguyên tệ', 'Có nguyên tệ', 'Chênh lệch (Nợ - Có)']
        error_fill = PatternFill(start_color='4338ca', end_color='4338ca', fill_type='solid')
        for col, h in enumerate(disc_headers, 1):
            cell = ws2.cell(row=start_row + 1, column=col, value=h)
            cell.font = header_font
            cell.fill = error_fill
            cell.alignment = header_align
            cell.border = thin_border
        ws2.row_dimensions[start_row + 1].height = 30

        for i, (_, row) in enumerate(df_discrepancy.iterrows(), start_row + 2):
            for j, col_name in enumerate(disc_headers, 1):
                val = row.get(col_name, '')
                cell = ws2.cell(row=i, column=j, value=val)
                cell.font = cell_font
                cell.border = thin_border
                if col_name in ('Nợ nguyên tệ', 'Có nguyên tệ', 'Chênh lệch (Nợ - Có)'):
                    cell.alignment = number_align
                    cell.number_format = '#,##0'
                else:
                    cell.alignment = cell_align

        # Totals for discrepancy
        t_row = start_row + 2 + len(df_discrepancy)
        ws2.merge_cells(f'A{t_row}:F{t_row}')
        ws2.cell(row=t_row, column=1, value='TỔNG CỘNG').font = Font(name='Times New Roman', bold=True, size=11)
        ws2.cell(row=t_row, column=1).alignment = Alignment(horizontal='center')
        ws2.cell(row=t_row, column=1).border = thin_border
        for c in range(2, 7):
            ws2.cell(row=t_row, column=c).border = thin_border

        for col_idx, col_name in [(7, 'Nợ nguyên tệ'), (8, 'Có nguyên tệ'), (9, 'Chênh lệch (Nợ - Có)')]:
            val = df_discrepancy[col_name].sum() if col_name in df_discrepancy.columns else 0
            cell = ws2.cell(row=t_row, column=col_idx, value=val)
            cell.font = Font(name='Times New Roman', bold=True, size=11)
            cell.number_format = '#,##0'
            cell.alignment = number_align
            cell.border = thin_border
    else:
        ws2.merge_cells('A11:J11')
        ws2['A11'] = '✓ Không có chênh lệch - Tổng phát sinh Nợ = Tổng phát sinh Có'
        ws2['A11'].font = Font(name='Times New Roman', bold=True, size=12, color='059669')
        ws2['A11'].alignment = Alignment(horizontal='center')

    # Column widths for sheet 2
    widths2 = [6, 10, 22, 18, 12, 45, 20, 20, 20]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(output)
    return output.getvalue()


# ═══════════════════════════════════════════════════
#                    MAIN APP
# ═══════════════════════════════════════════════════

# ─── Hero Header ───
st.markdown("""
<div class="hero-header">
    <h1>📊 Đối Chiếu Trung Gian</h1>
    <p>Công cụ phân tích & đối chiếu sổ chi tiết tài khoản — Kiểm tra cân đối Nợ / Có</p>
</div>
""", unsafe_allow_html=True)

# ─── Init Session State ───
if 'df' not in st.session_state:
    st.session_state.df = None
if 'analyzed' not in st.session_state:
    st.session_state.analyzed = False
if 'results' not in st.session_state:
    st.session_state.results = None
if 'discrepancy_df' not in st.session_state:
    st.session_state.discrepancy_df = None
if 'total_vals' not in st.session_state:
    st.session_state.total_vals = None
if 'file_name' not in st.session_state:
    st.session_state.file_name = None

# ─── Step 1: Add File ───
st.markdown("""
<div class="glass-panel">
    <h3 style="color: #a5b4fc; margin-top:0;">📁 Bước 1: Thêm file Excel</h3>
</div>
""", unsafe_allow_html=True)

uploaded_file = st.file_uploader(
    "Kéo thả hoặc chọn file Excel (.xlsx, .xls)",
    type=['xlsx', 'xls'],
    key='file_uploader',
    help="Hỗ trợ file Sổ chi tiết tài khoản trung gian"
)

if uploaded_file is not None:
    st.session_state.file_name = uploaded_file.name

    # Parse file
    with st.spinner("🔄 Đang đọc file..."):
        df, total_vals, wb, sheet_name = parse_excel_file(uploaded_file)

    if df is not None:
        st.session_state.df = df
        st.session_state.total_vals = total_vals

        # Show file info
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.markdown(f"""
            <div class="stat-card">
                <div class="stat-label">📄 File</div>
                <div class="stat-value blue" style="font-size:0.9rem;">{uploaded_file.name}</div>
            </div>
            """, unsafe_allow_html=True)
        with col2:
            st.markdown(f"""
            <div class="stat-card">
                <div class="stat-label">📋 Sheet</div>
                <div class="stat-value blue" style="font-size:0.9rem;">{sheet_name}</div>
            </div>
            """, unsafe_allow_html=True)
        with col3:
            st.markdown(f"""
            <div class="stat-card">
                <div class="stat-label">📊 Số dòng dữ liệu</div>
                <div class="stat-value green">{len(df)}</div>
            </div>
            """, unsafe_allow_html=True)
        with col4:
            st.markdown(f"""
            <div class="stat-card">
                <div class="stat-label">💰 Tổng Nợ (file)</div>
                <div class="stat-value amber">{format_number(total_vals['Tổng Nợ (file)']) if total_vals else 'N/A'}</div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

        # Preview data
        with st.expander("👁️ Xem trước dữ liệu", expanded=False):
            display_df = df.copy()
            display_df['Nợ nguyên tệ'] = display_df['Nợ nguyên tệ'].apply(format_number)
            display_df['Có nguyên tệ'] = display_df['Có nguyên tệ'].apply(format_number)
            st.dataframe(display_df, use_container_width=True, height=400)

        # ─── Step 2: Analyze ───
        st.markdown("""
        <div class="glass-panel">
            <h3 style="color: #a5b4fc; margin-top:0;">🔍 Bước 2: Phân tích đối chiếu</h3>
        </div>
        """, unsafe_allow_html=True)

        col_btn1, col_btn2, _ = st.columns([1, 1, 3])
        with col_btn1:
            analyze_clicked = st.button("🔍 Phân tích", type="primary", use_container_width=True)

        if analyze_clicked:
            with st.spinner("⏳ Đang phân tích dữ liệu..."):
                results, discrepancy_df = analyze_data(df, total_vals)
                st.session_state.results = results
                st.session_state.discrepancy_df = discrepancy_df
                st.session_state.analyzed = True

        # ─── Show Results ───
        if st.session_state.analyzed and st.session_state.results:
            results = st.session_state.results
            discrepancy_df = st.session_state.discrepancy_df

            st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

            # Summary cards
            st.markdown("""
            <div class="glass-panel animate-in">
                <h3 style="color: #e0e7ff; margin-top:0;">📈 Kết quả phân tích</h3>
            </div>
            """, unsafe_allow_html=True)

            c1, c2, c3, c4 = st.columns(4)
            with c1:
                st.markdown(f"""
                <div class="stat-card animate-in">
                    <div class="stat-label">Tổng phát sinh Nợ</div>
                    <div class="stat-value blue">{format_number(results['total_no'])}</div>
                </div>
                """, unsafe_allow_html=True)
            with c2:
                st.markdown(f"""
                <div class="stat-card animate-in">
                    <div class="stat-label">Tổng phát sinh Có</div>
                    <div class="stat-value green">{format_number(results['total_co'])}</div>
                </div>
                """, unsafe_allow_html=True)
            with c3:
                st.markdown(f"""
                <div class="stat-card animate-in">
                    <div class="stat-label">Chênh lệch</div>
                    <div class="stat-value {'green' if results['is_balanced'] else 'red'}">{format_number(results['chenh_lech'])}</div>
                </div>
                """, unsafe_allow_html=True)
            with c4:
                st.markdown(f"""
                <div class="stat-card animate-in">
                    <div class="stat-label">Trạng thái</div>
                    <div style="margin-top:0.5rem;">
                        <span class="{'badge-ok' if results['is_balanced'] else 'badge-error'}">
                            {'✓ Cân bằng' if results['is_balanced'] else '✗ Chênh lệch'}
                        </span>
                    </div>
                </div>
                """, unsafe_allow_html=True)

            st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

            # ──── Person Summary Table ────
            person_summary_df = results.get('person_summary_df')
            if person_summary_df is not None and not person_summary_df.empty:
                st.markdown("""
                <div class="glass-panel animate-in">
                    <h3 style="color: #e0e7ff; margin-top:0;">👥 Tổng hợp theo Người tạo</h3>
                    <p style="color: rgba(255,255,255,0.4); font-size: 0.8rem; margin-bottom:0.5rem;">
                        Nhóm theo người tạo, tính tổng Nợ - Có. Người tạo cân bằng (chênh lệch = 0) sẽ được bỏ qua.
                    </p>
                </div>
                """, unsafe_allow_html=True)

                # Build HTML table for person summary
                balanced_count = len(person_summary_df[person_summary_df['Cân bằng'] == True])
                unbalanced_count = len(person_summary_df[person_summary_df['Cân bằng'] == False])

                c_a, c_b = st.columns(2)
                with c_a:
                    st.markdown(f"""
                    <div class="stat-card" style="border-color: rgba(52,211,153,0.2);">
                        <div class="stat-label">✅ Người tạo cân bằng (bỏ qua)</div>
                        <div class="stat-value green">{balanced_count}</div>
                    </div>
                    """, unsafe_allow_html=True)
                with c_b:
                    st.markdown(f"""
                    <div class="stat-card" style="border-color: rgba(248,113,113,0.2);">
                        <div class="stat-label">⚠️ Người tạo có chênh lệch</div>
                        <div class="stat-value red">{unbalanced_count}</div>
                    </div>
                    """, unsafe_allow_html=True)

                # Show full person summary in expander
                with st.expander("📋 Xem chi tiết tổng hợp theo Người tạo", expanded=False):
                    display_person = person_summary_df.copy()
                    for nc in ['Tổng Nợ', 'Tổng Có', 'Chênh lệch']:
                        display_person[nc] = display_person[nc].apply(format_number)
                    display_person['Trạng thái'] = display_person['Cân bằng'].apply(
                        lambda x: '✅ Cân bằng' if x else '❌ Chênh lệch'
                    )
                    display_person = display_person.drop(columns=['Cân bằng'])
                    st.dataframe(display_person, use_container_width=True, height=300)

            st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

            # ──── Smart Matching Results ────
            matched_details = results.get('matched_details', {})
            unbalanced_persons = results.get('unbalanced_persons', [])

            if unbalanced_persons and matched_details:
                st.markdown("""
                <div class="glass-panel animate-in">
                    <h3 style="color: #fbbf24; margin-top:0;">🔗 Kết quả đối chiếu thông minh</h3>
                    <p style="color: rgba(255,255,255,0.4); font-size: 0.8rem;">
                        Với mỗi người tạo có chênh lệch, hệ thống so khớp: <br>
                        ① Nợ = Có + diễn giải giống ≥80% → bỏ qua<br>
                        ② Nợ dương + Nợ âm triệt tiêu + diễn giải giống ≥80% → bỏ qua<br>
                        ③ Giữ lại các dòng không xác định được
                    </p>
                </div>
                """, unsafe_allow_html=True)

                # Matching stats per person
                stats_html = '<div class="glass-panel" style="border-color: rgba(251,191,36,0.15);">'
                stats_html += '<table class="diff-table"><thead><tr>'
                stats_html += '<th>Người tạo</th><th style="text-align:center">Tổng dòng</th>'
                stats_html += '<th style="text-align:center">Đã khớp (bỏ qua)</th>'
                stats_html += '<th style="text-align:center">Chưa xác định</th>'
                stats_html += '</tr></thead><tbody>'

                total_matched = 0
                total_unmatched = 0
                for person in unbalanced_persons:
                    detail = matched_details.get(person, {})
                    total_r = detail.get('total_rows', 0)
                    matched_r = detail.get('matched', 0)
                    unmatched_r = detail.get('unmatched', 0)
                    total_matched += matched_r
                    total_unmatched += unmatched_r
                    stats_html += f'<tr>'
                    stats_html += f'<td>{person}</td>'
                    stats_html += f'<td style="text-align:center">{total_r}</td>'
                    stats_html += f'<td style="text-align:center; color:#34d399; font-weight:600;">{matched_r}</td>'
                    stats_html += f'<td style="text-align:center; color:#f87171; font-weight:600;">{unmatched_r}</td>'
                    stats_html += f'</tr>'

                stats_html += f'<tr style="border-top: 2px solid rgba(255,255,255,0.1);">'
                stats_html += f'<td><strong>TỔNG CỘNG</strong></td>'
                stats_html += f'<td style="text-align:center; font-weight:700;">{total_matched + total_unmatched}</td>'
                stats_html += f'<td style="text-align:center; color:#34d399; font-weight:700;">{total_matched}</td>'
                stats_html += f'<td style="text-align:center; color:#f87171; font-weight:700;">{total_unmatched}</td>'
                stats_html += f'</tr>'
                stats_html += '</tbody></table></div>'

                st.markdown(stats_html, unsafe_allow_html=True)

            st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

            # ──── Unmatched Rows Detail ────
            if discrepancy_df is not None and not discrepancy_df.empty:
                st.markdown(f"""
                <div class="glass-panel animate-in">
                    <h3 style="color: #f87171; margin-top:0;">⚠️ Các dòng chưa xác định gây chênh lệch ({len(discrepancy_df)} dòng)</h3>
                    <p style="color: rgba(255,255,255,0.5); font-size: 0.85rem;">
                        Các dòng dưới đây không thể khớp được qua phân tích thông minh — đây là nguyên nhân gây chênh lệch.
                    </p>
                </div>
                """, unsafe_allow_html=True)

                # Show per person
                if 'Người tạo' in discrepancy_df.columns:
                    for person in discrepancy_df['Người tạo'].unique():
                        person_disc = discrepancy_df[discrepancy_df['Người tạo'] == person]
                        person_sum_no = person_disc['Nợ nguyên tệ'].sum()
                        person_sum_co = person_disc['Có nguyên tệ'].sum()
                        person_cl = person_sum_no - person_sum_co

                        st.markdown(f"""
                        <div style="background: rgba(99,102,241,0.08); border-left: 3px solid #818cf8;
                                    padding: 0.6rem 1rem; border-radius: 0 8px 8px 0; margin-bottom: 0.5rem;">
                            <span style="color: #a5b4fc; font-weight: 700;">👤 {person}</span>
                            <span style="color: rgba(255,255,255,0.4); margin-left: 1rem; font-size: 0.8rem;">
                                {len(person_disc)} dòng chưa xác định | Chênh lệch: <span style="color:#f87171; font-weight:600;">{format_number(person_cl)}</span>
                            </span>
                        </div>
                        """, unsafe_allow_html=True)

                        display_disc = person_disc.copy()
                        display_cols = ['STT', 'Dòng Excel', 'Nguồn phát sinh', 'Ngày giao dịch', 'Số giao dịch',
                                        'Nội dung', 'Nợ nguyên tệ', 'Có nguyên tệ', 'Chênh lệch (Nợ - Có)']
                        display_disc = display_disc[[c for c in display_cols if c in display_disc.columns]]

                        for num_col in ['Nợ nguyên tệ', 'Có nguyên tệ', 'Chênh lệch (Nợ - Có)']:
                            if num_col in display_disc.columns:
                                display_disc[num_col] = display_disc[num_col].apply(format_number)

                        st.dataframe(
                            display_disc,
                            use_container_width=True,
                            height=min(400, 50 + len(display_disc) * 35),
                            column_config={
                                'STT': st.column_config.NumberColumn('STT', width='small'),
                                'Dòng Excel': st.column_config.NumberColumn('Dòng Excel', width='small'),
                                'Nội dung': st.column_config.TextColumn('Nội dung', width='large'),
                            }
                        )

                # Overall summary
                sum_no = discrepancy_df['Nợ nguyên tệ'].sum()
                sum_co = discrepancy_df['Có nguyên tệ'].sum()
                sum_cl = discrepancy_df['Chênh lệch (Nợ - Có)'].sum()

                st.markdown(f"""
                <div class="glass-panel" style="border-color: rgba(248,113,113,0.2);">
                    <table class="diff-table" style="max-width: 600px; margin: 0 auto;">
                        <thead>
                            <tr>
                                <th>Chỉ tiêu</th>
                                <th style="text-align:right">Giá trị</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td>Tổng Nợ (dòng chưa xác định)</td>
                                <td style="text-align:right; color:#818cf8; font-weight:600;">{format_number(sum_no)}</td>
                            </tr>
                            <tr>
                                <td>Tổng Có (dòng chưa xác định)</td>
                                <td style="text-align:right; color:#34d399; font-weight:600;">{format_number(sum_co)}</td>
                            </tr>
                            <tr style="border-top: 2px solid rgba(248,113,113,0.3);">
                                <td><strong>Tổng chênh lệch</strong></td>
                                <td style="text-align:right; color:#f87171; font-weight:700; font-size:1.1rem;">{format_number(sum_cl)}</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
                """, unsafe_allow_html=True)

            elif results['is_balanced']:
                st.markdown("""
                <div class="glass-panel animate-in" style="text-align:center; border-color: rgba(52,211,153,0.3);">
                    <div style="font-size: 3rem; margin-bottom: 0.5rem;">✅</div>
                    <h3 style="color: #34d399; margin:0;">Tổng phát sinh Nợ = Tổng phát sinh Có</h3>
                    <p style="color: rgba(255,255,255,0.5); margin-top:0.5rem;">Dữ liệu cân đối, không có chênh lệch.</p>
                </div>
                """, unsafe_allow_html=True)
            else:
                # All unbalanced rows were matched through smart matching
                st.markdown("""
                <div class="glass-panel animate-in" style="text-align:center; border-color: rgba(251,191,36,0.3);">
                    <div style="font-size: 3rem; margin-bottom: 0.5rem;">🔍</div>
                    <h3 style="color: #fbbf24; margin:0;">Tất cả các dòng đã được khớp thông minh</h3>
                    <p style="color: rgba(255,255,255,0.5); margin-top:0.5rem;">
                        Mặc dù tổng có chênh lệch, nhưng tất cả các dòng đã được giải thích qua đối chiếu thông minh.
                    </p>
                </div>
                """, unsafe_allow_html=True)

            # ─── Step 3: Export ───
            st.markdown("""
            <div class="glass-panel">
                <h3 style="color: #a5b4fc; margin-top:0;">📥 Bước 3: Xuất file báo cáo</h3>
                <p style="color: rgba(255,255,255,0.5); font-size: 0.85rem; margin-bottom:0;">
                    File xuất ra gồm 2 sheet: <strong>Dữ liệu gốc</strong> và <strong>Báo cáo chênh lệch</strong>
                </p>
            </div>
            """, unsafe_allow_html=True)

            export_data = create_export(df, discrepancy_df, results)
            export_filename = f"BaoCao_DoiChieu_{st.session_state.file_name.rsplit('.', 1)[0]}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

            st.download_button(
                label="📥 Xuất file báo cáo Excel",
                data=export_data,
                file_name=export_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

else:
    # Empty state
    st.markdown("""
    <div class="glass-panel" style="text-align:center; padding: 3rem;">
        <div style="font-size: 4rem; margin-bottom: 1rem; opacity: 0.3;">📂</div>
        <h3 style="color: rgba(255,255,255,0.4); margin:0;">Chưa có file nào được tải lên</h3>
        <p style="color: rgba(255,255,255,0.2); margin-top:0.5rem;">Vui lòng tải lên file Excel sổ chi tiết tài khoản để bắt đầu phân tích</p>
    </div>
    """, unsafe_allow_html=True)

# ─── Footer ───
st.markdown("""
<div style="text-align:center; padding: 2rem 0 1rem; color: rgba(255,255,255,0.15); font-size: 0.75rem;">
    Đối Chiếu Trung Gian © 2026 — Công cụ phân tích sổ chi tiết tài khoản
</div>
""", unsafe_allow_html=True)
