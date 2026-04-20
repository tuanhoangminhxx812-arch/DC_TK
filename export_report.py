"""Script xuất báo cáo đối chiếu trung gian - Phiên bản cải tiến v3 (khớp nhóm)."""
import pandas as pd
import numpy as np
from datetime import datetime
from itertools import combinations
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os

INPUT_FILE = r"D:\HOC A.I\đối chiếu trung gian\GL_016_33193.xlsx"
OUTPUT_DIR = r"D:\HOC A.I\đối chiếu trung gian"


# ═══ ĐỌC FILE ═══
print("Đang đọc file...")
wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
ws = wb.active
is_new_format = False
header_row = data_start = data_end = None

for i, row in enumerate(ws.iter_rows(min_row=1), 1):
    vals = [c.value for c in row]
    v0 = str(vals[0]).lower().strip() if vals[0] else ""
    if "nguồn phát sinh" in v0:
        is_new_format = True; header_row = i;
        if data_start is None: data_start = i + 1
    elif "nguồn bút toán" in v0: header_row = i
    if "dư đầu kỳ" in v0: data_start = i + 1
    if "tổng cộng" in v0 or "cộng phát sinh" in v0: data_end = i - 1; break

data_rows = []
if data_start is not None and data_end is None: data_end = ws.max_row
if data_start and data_end:
    for i, row in enumerate(ws.iter_rows(min_row=data_start, max_row=data_end), data_start):
        vals = [c.value for c in row]
        vals.extend([None] * (25 - len(vals)))
        if is_new_format:
            if vals[0] is None and vals[1] is None and vals[13] is None and vals[14] is None: continue
            data_rows.append({
                "STT": len(data_rows)+1, "Dòng Excel": i,
                "Nguồn phát sinh": vals[0] or "",
                "Số giao dịch": str(vals[1]).lstrip("'") if vals[1] else "",
                "Ngày giao dịch": vals[2].strftime("%d/%m/%Y") if hasattr(vals[2],"strftime") else (str(vals[2]) if vals[2] else ""),
                "Nội dung": vals[17] or "",
                "Nợ nguyên tệ": vals[13] if vals[13] is not None else 0,
                "Có nguyên tệ": vals[14] if vals[14] is not None else 0,
                "Người tạo": str(vals[19]).strip() if vals[19] else "",
            })
        else:
            if vals[0] is None and vals[1] is None and vals[5] is None and vals[6] is None: continue
            data_rows.append({
                "STT": len(data_rows)+1, "Dòng Excel": i,
                "Nguồn phát sinh": vals[0] or "",
                "Số giao dịch": str(vals[2]).lstrip("'") if vals[2] else "",
                "Ngày giao dịch": vals[1].strftime("%d/%m/%Y") if hasattr(vals[1],"strftime") else (str(vals[1]) if vals[1] else ""),
                "Nội dung": vals[4] or "",
                "Nợ nguyên tệ": vals[5] if vals[5] is not None else 0,
                "Có nguyên tệ": vals[6] if vals[6] is not None else 0,
                "Người tạo": str(vals[7]).strip() if vals[7] else "",
            })

df = pd.DataFrame(data_rows)
total_no = df["Nợ nguyên tệ"].sum()
total_co = df["Có nguyên tệ"].sum()
chenh_lech = total_no - total_co
print(f"  Tổng dòng: {len(df)} | Nợ: {total_no:,.0f} | Có: {total_co:,.0f} | CL: {chenh_lech:,.0f}")

# ═══ PHÂN TÍCH ═══
print("\n══ Bước 1: Nhóm theo Người tạo ══")
unbalanced = []
for person, gdf in df.groupby("Người tạo", sort=False):
    cl = gdf["Nợ nguyên tệ"].sum() - gdf["Có nguyên tệ"].sum()
    if abs(cl) >= 0.01: unbalanced.append(person)
print(f"  Người tạo có chênh lệch: {len(unbalanced)}")

print("\n══ Bước 2: Khớp trong cùng người tạo ══")
global_matched = set()
for person in unbalanced:
    pdf = df[df["Người tạo"]==person].copy().reset_index(drop=True)
    matched = [False]*len(pdf)
    no_rows, co_rows = [], []
    for idx, row in pdf.iterrows():
        nv = row["Nợ nguyên tệ"] if pd.notna(row["Nợ nguyên tệ"]) else 0
        cv = row["Có nguyên tệ"] if pd.notna(row["Có nguyên tệ"]) else 0
        if abs(nv)>0 and abs(cv)==0: no_rows.append(idx)
        elif abs(cv)>0 and abs(nv)==0: co_rows.append(idx)

    # Pass 1: Nợ = Có
    used_co = set()
    for ni in no_rows:
        if matched[ni]: continue
        nv = pdf.loc[ni,"Nợ nguyên tệ"]
        for ci in co_rows:
            if ci in used_co or matched[ci]: continue
            cv = pdf.loc[ci,"Có nguyên tệ"]
            if abs(abs(nv)-abs(cv))<0.01: matched[ni]=matched[ci]=True; used_co.add(ci); break

    # Pass 2: Nợ+ + Nợ-
    umi = [i for i in no_rows if not matched[i]]
    pos_no = [i for i in umi if pdf.loc[i,"Nợ nguyên tệ"]>0]
    neg_no = [i for i in umi if pdf.loc[i,"Nợ nguyên tệ"]<0]
    used_neg = set()
    for pi in pos_no:
        if matched[pi]: continue
        pv = pdf.loc[pi,"Nợ nguyên tệ"]
        for ngi in neg_no:
            if ngi in used_neg or matched[ngi]: continue
            nv = pdf.loc[ngi,"Nợ nguyên tệ"]
            if abs(pv+nv)<0.01: matched[pi]=matched[ngi]=True; used_neg.add(ngi); break

    # Pass 2b: Nhóm Nợ triệt tiêu (tổng = 0)
    still_unmatched_no = [i for i in no_rows if not matched[i]]
    if len(still_unmatched_no) > 1:
        no_sum = sum(pdf.loc[i,"Nợ nguyên tệ"] for i in still_unmatched_no)
        if abs(no_sum) < 0.01:
            for i in still_unmatched_no: matched[i] = True

    # Pass 3: Có+ + Có-
    uci = [i for i in co_rows if not matched[i]]
    pos_co = [i for i in uci if pdf.loc[i,"Có nguyên tệ"]>0]
    neg_co = [i for i in uci if pdf.loc[i,"Có nguyên tệ"]<0]
    used_neg_co = set()
    for pi in pos_co:
        if matched[pi]: continue
        pv = pdf.loc[pi,"Có nguyên tệ"]
        for ngi in neg_co:
            if ngi in used_neg_co or matched[ngi]: continue
            nv = pdf.loc[ngi,"Có nguyên tệ"]
            if abs(pv+nv)<0.01: matched[pi]=matched[ngi]=True; used_neg_co.add(ngi); break

    # Pass 3b: Nhóm Có triệt tiêu (tổng = 0)
    still_unmatched_co = [i for i in co_rows if not matched[i]]
    if len(still_unmatched_co) > 1:
        co_sum = sum(pdf.loc[i,"Có nguyên tệ"] for i in still_unmatched_co)
        if abs(co_sum) < 0.01:
            for i in still_unmatched_co: matched[i] = True

    for idx in range(len(pdf)):
        if matched[idx]: global_matched.add(pdf.loc[idx,"STT"])
    mc = sum(matched)
    print(f"  {person}: {len(pdf)} dòng, khớp {mc}, còn {len(pdf)-mc}")

# Pass 4: Khớp chéo 1-1
print("\n══ Bước 3: Khớp chéo 1-1 giữa người tạo ══")
unmatched_df = df[(df["Người tạo"].isin(unbalanced))&(~df["STT"].isin(global_matched))].copy()
cross_no = unmatched_df[(unmatched_df["Nợ nguyên tệ"].abs()>0)&((unmatched_df["Có nguyên tệ"].fillna(0)).abs()==0)]
cross_co = unmatched_df[(unmatched_df["Có nguyên tệ"].abs()>0)&((unmatched_df["Nợ nguyên tệ"].fillna(0)).abs()==0)]
cross_stts = set()
used = set()
for _, nr in cross_no.iterrows():
    if nr["STT"] in cross_stts: continue
    for _, cr in cross_co.iterrows():
        if cr["STT"] in used or cr["STT"] in cross_stts: continue
        if abs(abs(nr["Nợ nguyên tệ"])-abs(cr["Có nguyên tệ"]))<0.01:
            cross_stts.add(nr["STT"]); cross_stts.add(cr["STT"]); used.add(cr["STT"]); break
global_matched.update(cross_stts)
print(f"  Khớp chéo 1-1: {len(cross_stts)} dòng")

# Pass 5: Khớp NHÓM
print("\n══ Bước 4: Khớp nhóm (tổng nhiều Nợ = tổng nhiều Có) ══")
remaining_df = df[(df["Người tạo"].isin(unbalanced))&(~df["STT"].isin(global_matched))].copy()
group_matched = set()

if len(remaining_df) > 0:
    rem_no = remaining_df[remaining_df["Nợ nguyên tệ"].fillna(0).abs()>0]
    rem_co = remaining_df[(remaining_df["Có nguyên tệ"].fillna(0).abs()>0)&(remaining_df["Nợ nguyên tệ"].fillna(0).abs()==0)]

    no_items = [(r["STT"], r["Nợ nguyên tệ"]) for _, r in rem_no.iterrows() if abs(r["Nợ nguyên tệ"])>0]
    co_items = [(r["STT"], r["Có nguyên tệ"]) for _, r in rem_co.iterrows() if abs(r["Có nguyên tệ"])>0]

    if no_items and co_items:
        total_rem_no = sum(v for _,v in no_items)
        total_rem_co = sum(v for _,v in co_items)

        if abs(total_rem_no - total_rem_co) < 0.01:
            print(f"  Tổng Nợ chưa khớp = Tổng Có chưa khớp = {total_rem_no:,.0f}")
            print(f"  → Tất cả {len(no_items)+len(co_items)} dòng khớp nhóm!")
            for s,_ in no_items: group_matched.add(s)
            for s,_ in co_items: group_matched.add(s)
        else:
            # Tìm subset khớp
            max_co_combo = min(len(co_items), 5)
            co_sum_map = {}
            for size in range(1, max_co_combo+1):
                for combo in combinations(co_items, size):
                    cs = round(sum(v for _,v in combo), 2)
                    cstts = frozenset(s for s,_ in combo)
                    if cs not in co_sum_map: co_sum_map[cs] = []
                    co_sum_map[cs].append(cstts)

            max_no_combo = min(len(no_items), 5)
            for size in range(1, max_no_combo+1):
                for combo in combinations(no_items, size):
                    cs = round(sum(v for _,v in combo), 2)
                    if cs in co_sum_map:
                        nstts = frozenset(s for s,_ in combo)
                        if not nstts.intersection(group_matched):
                            for costts in co_sum_map[cs]:
                                if not costts.intersection(group_matched):
                                    print(f"  Khớp: Nợ STT {sorted(nstts)} = Có STT {sorted(costts)}")
                                    group_matched.update(nstts)
                                    group_matched.update(costts)
                                    break

global_matched.update(group_matched)
print(f"  Tổng khớp nhóm: {len(group_matched)} dòng")

# Kết quả cuối cùng
all_unmatched = []
for _, row in df.iterrows():
    if row["Người tạo"] in unbalanced and row["STT"] not in global_matched:
        nv = row["Nợ nguyên tệ"] if pd.notna(row["Nợ nguyên tệ"]) else 0
        cv = row["Có nguyên tệ"] if pd.notna(row["Có nguyên tệ"]) else 0
        rd = row.to_dict()
        rd["Chênh lệch (Nợ - Có)"] = nv - cv
        all_unmatched.append(rd)

disc_df = pd.DataFrame(all_unmatched) if all_unmatched else pd.DataFrame()
print(f"\n{'='*50}")
print(f"KẾT QUẢ CUỐI CÙNG: Còn {len(disc_df)} dòng chưa xác định")
print(f"{'='*50}")

if not disc_df.empty:
    for person in disc_df["Người tạo"].unique():
        pr = disc_df[disc_df["Người tạo"]==person]
        sn = pr["Nợ nguyên tệ"].sum()
        sc = pr["Có nguyên tệ"].sum()
        print(f"\n  {person}: {len(pr)} dòng | Nợ: {sn:,.0f} | Có: {sc:,.0f} | CL: {sn-sc:,.0f}")
        for _, r in pr.iterrows():
            nstr = f"Nợ:{r['Nợ nguyên tệ']:>14,.0f}" if r["Nợ nguyên tệ"]!=0 else "                  "
            cstr = f"Có:{r['Có nguyên tệ']:>14,.0f}" if r["Có nguyên tệ"]!=0 else "                  "
            print(f"    STT {r['STT']:>3} | {nstr} | {cstr} | {str(r['Nội dung'])[:55]}")

# ═══ XUẤT FILE ═══
print("\nĐang xuất file...")
out_wb = openpyxl.Workbook()
hf = Font(name="Times New Roman", bold=True, size=11, color="FFFFFF")
hfill = PatternFill(start_color="312e81", end_color="312e81", fill_type="solid")
ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
cf = Font(name="Times New Roman", size=10)
ca = Alignment(vertical="center", wrap_text=True)
na = Alignment(horizontal="right", vertical="center")
tb = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))

# Sheet 1
ws1 = out_wb.active; ws1.title = "Dữ liệu gốc"
ws1.merge_cells("A1:J1")
ws1["A1"] = "DỮ LIỆU GỐC - SỔ CHI TIẾT TÀI KHOẢN"
ws1["A1"].font = Font(name="Times New Roman", bold=True, size=14, color="312e81")
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
now_str = datetime.now().strftime("%d/%m/%Y %H:%M")
ws1.merge_cells("A2:J2")
ws1["A2"] = "Ngày xuất: " + now_str
ws1["A2"].font = Font(name="Times New Roman", italic=True, size=9, color="666666")
ws1["A2"].alignment = Alignment(horizontal="center")

headers1 = ["STT","Dòng Excel","Nguồn phát sinh", "Số giao dịch", "Ngày giao dịch","Nội dung","Nợ nguyên tệ","Có nguyên tệ","Người tạo"]
for col,h in enumerate(headers1,1):
    c=ws1.cell(row=4,column=col,value=h); c.font=hf; c.fill=hfill; c.alignment=ha; c.border=tb

for i,(_,row) in enumerate(df.iterrows(),5):
    for j,cn in enumerate(headers1,1):
        v=row.get(cn,""); c=ws1.cell(row=i,column=j,value=v); c.font=cf; c.border=tb
        if cn in ("Nợ nguyên tệ","Có nguyên tệ"): c.alignment=na; c.number_format="#,##0"
        else: c.alignment=ca

tr=5+len(df)
ws1.merge_cells(f"A{tr}:F{tr}")
c=ws1.cell(row=tr,column=1,value="TỔNG CỘNG"); c.font=Font(name="Times New Roman",bold=True,size=11); c.alignment=Alignment(horizontal="center"); c.border=tb
for col in range(2,7): ws1.cell(row=tr,column=col).border=tb
for ci,val in [(7,total_no),(8,total_co)]:
    c=ws1.cell(row=tr,column=ci,value=val); c.font=Font(name="Times New Roman",bold=True,size=11); c.number_format="#,##0"; c.alignment=na; c.border=tb
for i,w in enumerate([6,10,22,18,12,50,20,20,20],1): ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w

# Sheet 2
ws2 = out_wb.create_sheet(title="Báo cáo chênh lệch")
ws2.merge_cells("A1:K1")
ws2["A1"] = "BÁO CÁO KẾT QUẢ ĐỐI CHIẾU"; ws2["A1"].font=Font(name="Times New Roman",bold=True,size=14,color="312e81"); ws2["A1"].alignment=Alignment(horizontal="center",vertical="center")
ws2.merge_cells("A2:K2")
ws2["A2"] = "Ngày xuất: " + now_str; ws2["A2"].font=Font(name="Times New Roman",italic=True,size=9,color="666666"); ws2["A2"].alignment=Alignment(horizontal="center")

sf = PatternFill(start_color="f0f0ff",end_color="f0f0ff",fill_type="solid")
ws2.merge_cells("A4:C4"); ws2["A4"]="TÓM TẮT KẾT QUẢ"; ws2["A4"].font=Font(name="Times New Roman",bold=True,size=12,color="312e81")
for idx,(label,value) in enumerate([("Tổng phát sinh Nợ:",total_no),("Tổng phát sinh Có:",total_co),("Chênh lệch (Nợ - Có):",chenh_lech),("Kết luận:","CÂN BẰNG" if abs(chenh_lech)<0.01 else "CHÊNH LỆCH")],5):
    c1=ws2.cell(row=idx,column=1,value=label); c1.font=Font(name="Times New Roman",bold=True,size=11); c1.fill=sf; c1.border=tb
    ws2.merge_cells(f"A{idx}:B{idx}")
    c3=ws2.cell(row=idx,column=3,value=value); c3.font=Font(name="Times New Roman",bold=True,size=11); c3.fill=sf; c3.border=tb
    if isinstance(value,(int,float)): c3.number_format="#,##0"

if not disc_df.empty:
    sr=11
    ws2.merge_cells(f"A{sr}:J{sr}"); ws2[f"A{sr}"]="CÁC DÒNG CHƯA XÁC ĐỊNH GÂY CHÊNH LỆCH"; ws2[f"A{sr}"].font=Font(name="Times New Roman",bold=True,size=12,color="dc2626")
    dh=["STT","Dòng Excel","Nguồn phát sinh","Số giao dịch","Ngày giao dịch","Nội dung","Nợ nguyên tệ","Có nguyên tệ","Người tạo","Chênh lệch (Nợ - Có)"]
    ef=PatternFill(start_color="4338ca",end_color="4338ca",fill_type="solid")
    for col,h in enumerate(dh,1): c=ws2.cell(row=sr+1,column=col,value=h); c.font=hf; c.fill=ef; c.alignment=ha; c.border=tb

    for i,(_,row) in enumerate(disc_df.iterrows(),sr+2):
        for j,cn in enumerate(dh,1):
            v=row.get(cn,""); c=ws2.cell(row=i,column=j,value=v); c.font=cf; c.border=tb
            if cn in ("Nợ nguyên tệ","Có nguyên tệ","Chênh lệch (Nợ - Có)"): c.alignment=na; c.number_format="#,##0"
            else: c.alignment=ca

    ttr=sr+2+len(disc_df)
    ws2.merge_cells(f"A{ttr}:F{ttr}"); c=ws2.cell(row=ttr,column=1,value="TỔNG CỘNG"); c.font=Font(name="Times New Roman",bold=True,size=11); c.alignment=Alignment(horizontal="center"); c.border=tb
    for col in range(2,7): ws2.cell(row=ttr,column=col).border=tb
    for ci,cn in [(7,"Nợ nguyên tệ"),(8,"Có nguyên tệ"),(10,"Chênh lệch (Nợ - Có)")]:
        v=disc_df[cn].sum() if cn in disc_df.columns else 0
        c=ws2.cell(row=ttr,column=ci,value=v); c.font=Font(name="Times New Roman",bold=True,size=11); c.number_format="#,##0"; c.alignment=na; c.border=tb
else:
    ws2.merge_cells("A11:J11"); ws2["A11"]="Không có chênh lệch"; ws2["A11"].font=Font(name="Times New Roman",bold=True,size=12,color="059669"); ws2["A11"].alignment=Alignment(horizontal="center")

for i,w in enumerate([6,10,22,18,12,45,20,20,20,20],1): ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w

fname = "BaoCao_DoiChieu_GL_016_33193_" + datetime.now().strftime("%Y%m%d_%H%M") + ".xlsx"
fpath = os.path.join(OUTPUT_DIR, fname)
out_wb.save(fpath)
print(f"\n✅ Đã xuất: {fpath}")
